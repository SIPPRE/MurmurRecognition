import os
import numpy as np
import pandas as pd
import librosa
from scipy.signal import hilbert
import pywt
import shutil
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Ορισμός των μονοπατιών
data_path = r"C:\Users\aggel\OneDrive\Υπολογιστής\Cardiac_Project\data\the-circor-digiscope-phonocardiogram-dataset-1.0.3\training_data"
output_csv_with_murmur = r"C:\Users\aggel\OneDrive\Υπολογιστής\Cardiac_Project\output\features_with_murmur.xlsx"
output_csv_without_murmur = r"C:\Users\aggel\OneDrive\Υπολογιστής\Cardiac_Project\output\features_without_murmur.xlsx"
output_csv_unknown_murmur = r"C:\Users\aggel\OneDrive\Υπολογιστής\Cardiac_Project\output\features_unknown_murmur.xlsx"
output_csv_combined = r"C:\Users\aggel\OneDrive\Υπολογιστής\Cardiac_Project\output\features_combined.xlsx"
training_data_csv = r"C:\Users\aggel\OneDrive\Υπολογιστής\Cardiac_Project\data\the-circor-digiscope-phonocardiogram-dataset-1.0.3\training_data.csv"

# Διαγραφή όλων των αρχείων στον φάκελο output
output_folder = r"C:\Users\aggel\OneDrive\Υπολογιστής\Cardiac_Project\output"
for filename in os.listdir(output_folder):
    file_path = os.path.join(output_folder, filename)
    try:
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
        elif os.path.isdir(file_path):
            shutil.rmtree(file_path)
    except Exception as e:
        print(f'Failed to delete {file_path}. Reason: {e}')

# Φόρτωση των δεδομένων εκπαίδευσης για φιλτράρισμα
training_data = pd.read_csv(training_data_csv)

# Δημιουργία dictionary για τις κατηγορίες
patient_categories = {row['Patient ID']: 1 if row['Outcome'] == 'Abnormal' else 0 for index, row in training_data.iterrows()}

# Φιλτράρισμα των αρχείων με φύσημα καρδιάς
patient_ids_with_murmur = training_data[training_data['Murmur'] == 'Present']['Patient ID'].astype(str).values
files_with_murmur = [f"{patient_id}_{loc}.wav" for patient_id in patient_ids_with_murmur for loc in ['AV', 'PV', 'TV', 'MV']]

# Φιλτράρισμα των αρχείων χωρίς φύσημα καρδιάς
patient_ids_without_murmur = training_data[training_data['Murmur'] == 'Absent']['Patient ID'].astype(str).values

# Φιλτράρισμα των αρχείων με άγνωστη κατάσταση
patient_ids_unknown_murmur = training_data[training_data['Murmur'] == 'Unknown']['Patient ID'].astype(str).values

print("Files with murmurs:")
for file in files_with_murmur:
    print(file)

# Συνάρτηση για την εξαγωγή χαρακτηριστικών από ένα αρχείο ήχου
def extract_features(file_name):
    y, sr = librosa.load(file_name, sr=4000, duration=30)
    mfccs = np.mean(librosa.feature.mfcc(y=y, sr=sr, n_mfcc=13).T, axis=0)
    features = {
        'filename': file_name,
        'spectral_centroid': np.mean(librosa.feature.spectral_centroid(y=y, sr=sr)),
        'zero_crossing_rate': np.mean(librosa.feature.zero_crossing_rate(y)),
        'rmse': np.mean(librosa.feature.rms(y=y)),
        'spectral_bandwidth': np.mean(librosa.feature.spectral_bandwidth(y=y, sr=sr)),
        'spectral_rolloff': np.mean(librosa.feature.spectral_rolloff(y=y, sr=sr)),
        'wavelet_entropy': wavelet_entropy(y, sr),
        'hilbert_huang_transform': hilbert_huang_transform(y, sr),
        'cwt': cwt_features(y, sr),
        'wst': wst_features(y, sr),
        'shannon_entropy': shannon_entropy(y),
        'signal_energy': signal_energy(y),
        'stft': stft_features(y, sr),
        'signal_envelope': signal_envelope(y),
        'teager_energy_operator': teager_energy_operator(y)
    }
    for i, mfcc in enumerate(mfccs):
        features[f'mfcc_{i+1}'] = mfcc
    return features

# Συνάρτηση για την εξαγωγή wavelet entropy
def wavelet_entropy(y, sr):
    coeffs = pywt.wavedec(y, 'db1')
    entropy = -np.sum([np.sum(c**2) * np.log(np.sum(c**2)) for c in coeffs])
    return entropy

# Συνάρτηση για την εξαγωγή Hilbert-Huang transform
def hilbert_huang_transform(y, sr):
    analytic_signal = hilbert(y)
    amplitude_envelope = np.abs(analytic_signal)
    instantaneous_phase = np.unwrap(np.angle(analytic_signal))
    instantaneous_frequency = np.diff(instantaneous_phase) / (2.0*np.pi) * sr
    return np.mean(instantaneous_frequency)

# Συνάρτηση για την εξαγωγή CWT features
def cwt_features(y, sr):
    widths = np.arange(1, 31)
    cwtmatr = pywt.cwt(y, widths, 'mexh')[0]
    return np.mean(cwtmatr)

# Συνάρτηση για την εξαγωγή WST features
def wst_features(y, sr):
    scattering = pywt.wavedec(y, 'db1', level=5)
    return np.mean(scattering[0])

# Συνάρτηση για την εξαγωγή Shannon entropy
def shannon_entropy(y):
    prob_y = y / np.sum(y)
    prob_y = prob_y[prob_y > 0]  # Φιλτράρουμε τις μηδενικές τιμές
    entropy = -np.sum(prob_y * np.log2(prob_y))
    return entropy

# Συνάρτηση για την εξαγωγή signal energy
def signal_energy(y):
    return np.sum(y**2)

# Συνάρτηση για την εξαγωγή STFT features
def stft_features(y, sr):
    stft = np.abs(librosa.stft(y))
    return np.mean(stft)

# Συνάρτηση για την εξαγωγή signal envelope
def signal_envelope(y):
    analytic_signal = hilbert(y)
    amplitude_envelope = np.abs(analytic_signal)
    return np.mean(amplitude_envelope)

# Συνάρτηση για την εξαγωγή Teager Energy Operator
def teager_energy_operator(y):
    teager = np.diff(y)**2 - y[:-1]*y[1:]
    return np.mean(teager)

# Εξαγωγή χαρακτηριστικών από όλα τα αρχεία ήχου και ενημέρωση του υπάρχοντος αρχείου CSV
def main():
    data_with_murmur = []
    data_without_murmur = []
    data_unknown_murmur = []

    files = sorted([file for file in os.listdir(data_path) if file.endswith('.wav')], key=lambda x: int(x.split('_')[0]))
    for file in tqdm(files, desc="Processing files"):
        file_id = file.split('_')[0]
        if file in files_with_murmur:
            features = extract_features(os.path.join(data_path, file))
            # Προσθήκη της κατηγορίας στο τέλος των χαρακτηριστικών
            features['category'] = 0  # 0 για φυσημα
            features['NormalOrAbnormal'] = patient_categories[int(file_id)]  # 0 για Normal, 1 για Abnormal
            data_with_murmur.append(features)
            print(f"Extracted features from {file} (with murmur)")
        elif file_id in patient_ids_without_murmur:
            features = extract_features(os.path.join(data_path, file))
            features['category'] = 1  # 1 για χωρις φυσημα
            features['NormalOrAbnormal'] = ''  # Κενή κατηγορία για ασθενείς χωρίς φύσημα
            data_without_murmur.append(features)
            print(f"Extracted features from {file} (without murmur)")
        elif file_id in patient_ids_unknown_murmur:
            data_unknown_murmur.append({'filename': file_id})
            print(f"File {file} is for a patient with unknown murmur status")

    df_with_murmur = pd.DataFrame(data_with_murmur)
    df_without_murmur = pd.DataFrame(data_without_murmur)
    df_unknown_murmur = pd.DataFrame(data_unknown_murmur)

    df_combined = pd.concat([df_with_murmur, df_without_murmur], ignore_index=True)

    def format_and_save(df, output_excel):
        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws["1:1"]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save(output_excel)
        print(f"Formatted {output_excel}")

    # Save the dataframes to their respective Excel files
    format_and_save(df_with_murmur, output_csv_with_murmur)
    format_and_save(df_without_murmur, output_csv_without_murmur)
    format_and_save(df_unknown_murmur, output_csv_unknown_murmur)
    format_and_save(df_combined, output_csv_combined)

    # Function to update filename column for the combined dataset
    def update_filename_column(output_excel):
        wb = load_workbook(output_excel)
        ws = wb.active

        # Create a dictionary for the mapping
        mapping = {'AV': 1, 'MV': 2, 'PV': 3, 'TV': 4}

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None:
                    for key, value in mapping.items():
                        if f"_{key}" in cell.value:
                            cell.value = value
                            break

        wb.save(output_excel)
        print(f"Updated {output_excel}")

    # Update filename columns for the combined dataset
    update_filename_column(output_csv_with_murmur)
    update_filename_column(output_csv_without_murmur)
    update_filename_column(output_csv_unknown_murmur)
    update_filename_column(output_csv_combined)

if __name__ == "__main__":
    main()

